from openpyxl import Workbook
from openpyxl import load_workbook
import json
import traceback

with open('config.json', 'r') as f:
    config = json.load(f)

class ImageRecord:
    def __init__(self, phone, contract, text):
        self.count = 1
        self.phone = phone
        self.contract = contract
        self.contents = []
        self.contents.append(text)
class ExcelProcessor:
    def __init__(self,file_name):
        self.header = ""
        self.file_name = file_name
    def process_file(self):
        try:
            wb = Workbook()
            wb = load_workbook(self.file_name, read_only=True)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]
            header_row = ws[1]
            ph_col = ws.max_column
            contact_row = {}
            max_row=config["max_row_in_iamge"]
            rows_in_image = 0
            self.header += "<tr>"
            for i in range(0,ph_col):
                if header_row[i].value == 'phone':
                    ph_col = i
                    break
                self.header += "<th>{}</th>".format(header_row[i].value)
            
            counter = 0
            for row in ws.rows:
                if counter == 0:
                    counter += 1
                    continue
                counter += 1
                print("processing row  = ", counter)
                if row[ph_col].value is None:
                    break
                text = "<tr>"
                for j in range(0,ph_col):
                    colorvalue = str(row[j].fill.start_color.value)[2:]
                    color = '#' + colorvalue if colorvalue != '000000' and len(colorvalue) == 6 else '#FFFFFF'
                    # if counter > 17 and counter < 25 and j == 0:
                        # print(str(counter),color,str(row[j].fill.start_color.value),sep=' | ')
                    text += ('<td bgcolor="{}" >{}</td>').format(color,str(row[j].value))
                text+="</tr>"
                phone_num = row[ph_col].value
                contract = str(+row[ph_col-1].value)
                phone_contract = phone_num + "_"+contract
                if  phone_contract in contact_row:
                    contact_row[phone_contract].count+=1
                    if rows_in_image < max_row:
                        contact_row[phone_contract].contents[-1]+=text
                        rows_in_image+=1
                    else:
                        contact_row[phone_contract].contents.append(text)
                        rows_in_image = 1
                else:
                    r = ImageRecord(phone_num,contract,text)
                    rows_in_image = 1
                    contact_row[phone_contract] = r
                print('processing row is finished')
            return contact_row
        except:
            traceback.print_exc()
            return None