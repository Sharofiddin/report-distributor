from telethon import TelegramClient, events, sync
from openpyxl import Workbook
from openpyxl import load_workbook
from telethon import errors
from openpyxl import styles
import time
import imgkit
import html2img
import json

with open('config.json', 'r') as f:
    config = json.load(f)

class Record:
    def __init__(self, phone, text):
        self.p = phone
        self.contents = []
        self.contents.append(text)
wb = Workbook()
wb = load_workbook(config["file_name"], read_only=True)
api_id = config["api_id"]
api_hash = config["api_hash"]
client = TelegramClient('session_name', api_id, api_hash)
client.start()
sheets = wb.sheetnames
ws = wb[sheets[0]]
header_row = ws[1]
ph_col = ws.max_column
contact_row = {}
max_row=config["max_row_in_iamge"]
rows_in_image = 0
header = "<tr>"
for i in range(0,ph_col):
    if header_row[i].value == 'phone':
        ph_col = i
        break
    header += "<th>{}</th>".format(header_row[i].value)
 
counter = 0
for row in ws.rows:
    if counter == 0:
        counter += 1
        continue
    counter += 1
    print("processing row  = ", counter)
    if row[ph_col].value is None:
        break
    text = "<tr>"
    for j in range(0,ph_col):
        color = '#'+str(row[j].fill.start_color.value)
        text += ('<td nowrap="nowrap" bgcolor="{}" >{}</td>').format(color,str(row[j].value))
    text+="</tr>"
    phone_num = row[ph_col].value
    phone_contract = phone_num + "_"+str(+row[ph_col-1].value)
    if  phone_contract in contact_row:
        if rows_in_image < 20:
            contact_row[phone_contract].contents[-1]+=text
            rows_in_image+=1
        else:
            contact_row[phone_contract].contents.append(text)
            rows_in_image = 1
    else:
        r = Record(phone_num,text)
        rows_in_image = 1
        contact_row[phone_contract] = r
for key in list(contact_row.keys()):
    for content in contact_row[key].contents:
        imgkit.from_string(html2img.prepare_body(header+content),'name.jpg')
        try:
            user = client.get_entity(contact_row[key].p)
            client.send_file(user, "name.jpg")
        except errors.FloodWaitError as e:
            print('Have to sleep', e.seconds, 'seconds')
            client.disconnect()
            client.start()
            user = client.get_entity(contact_row[key].p)
            client.send_file(user, "name.jpg")
